import logging
from datetime import datetime

from django.contrib.auth.decorators import login_required
from django.db.models import Q
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.utils.encoding import escape_uri_path
from django.views.decorators.http import require_http_methods
from django.views.generic import TemplateView, View
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook

from rest.models import Person, Record

tz = timezone.get_current_timezone()
logger = logging.getLogger(__name__)


@method_decorator(login_required, name='dispatch')
class ReportView(TemplateView):
    template_name = 'enhanced_ui/report.html'

    def get_context_data(self, **kwargs):
        date = kwargs.get('date')
        if date is None:
            today = datetime.now()
        else:
            today = datetime.strptime(date, '%Y-%m-%d')
        all_records = Record.objects.filter(created__date=today, record_type__in=['IN', 'OUT'])
        attended_pks = list(all_records.values_list('target', flat=True))

        attended_people = Person.objects.filter(Q(pk__in=attended_pks) & (Q(departure_date__gt=today) | Q(departure_date=None)) & Q(is_deleted=False))
        #attended_people = Person.objects.filter(Q(pk__in=attended_pks) & Q(is_active=True) & Q(is_deleted=False))
        attended_data = []
        for person in attended_people:
            starts = all_records.filter(Q(target=person) & Q(created__date=today) & Q(record_type='IN')).order_by(
                'created')
            start = starts.first().created.astimezone(tz).strftime('%Y-%m-%d %H:%M:%S') if starts.count() > 0 else '无记录'
            ends = all_records.filter(Q(target=person) & Q(created__date=today) & Q(record_type='OUT')).order_by(
                '-created')
            end = ends.first().created.astimezone(tz).strftime('%Y-%m-%d %H:%M:%S') if ends.count() > 0 else '无记录'
            attended_data.append({
                'pk': person.pk,
                'name': person.name,
                'position': person.position,
                'start': start,
                'end': end,
                'employee_number': person.employee_number,
                'department': person.department
            })

        absent_people = Person.objects.filter(Q(is_deleted=False) & (Q(departure_date__gt=today) | Q(departure_date=None))).exclude(pk__in=attended_pks)
        #absent_people = Person.objects.filter(Q(is_deleted=False) & (Q(is_active=True))).exclude(pk__in=attended_pks)

        context = super().get_context_data(**kwargs)
        context['date'] = today.strftime('%Y-%m-%d')
        context['absent_people'] = absent_people
        context['attended_people'] = attended_data
        return context


@method_decorator(login_required, name='dispatch')
class PersonalDateRangeReportView(TemplateView):
    template_name = 'enhanced_ui/personal_daterange_report.html'

    def get_context_data(self, **kwargs):
        logger.info('开始解析数据')
        person_pk = self.kwargs['pk']
        logger.debug('人员pk为{}'.format(person_pk))
        person = Person.objects.get(pk=person_pk)
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        attended_days, absent_days, additional_days, records = person.date_range_report(start_date, end_date)

        context = super().get_context_data(**kwargs)
        context['object'] = person
        context['start_date'] = start_date
        context['end_date'] = end_date
        context['attended_days'] = attended_days
        context['absent_days'] = absent_days
        context['additional_days'] = additional_days
        context['records'] = records
        return context


@require_http_methods(["GET"])
def get_personal_records_by_date(request):
    try:
        logger.info('开始解析数据')
        pk = request.GET.get('pk')
        date = request.GET.get('date')
        logger.debug('人员pk为{0}，查询日期为{1}'.format(pk, date))

        logger.info('开始查询人员{}'.format(pk))
        person = Person.objects.get(pk=pk)
        personal_information = {
            'name': person.name,
            'position': person.position
        }

        logger.info('开始过滤匹配记录')
        query_date = datetime.strptime(date, '%Y-%m-%d')
        records = Record.objects.filter(Q(target=person) & Q(created__date=query_date)).order_by('created')
        records = [{
            'created': record.created.astimezone(tz).strftime('%Y-%m-%d %H:%M:%S'),
            'record_type': record.record_type
        } for record in records]

        logger.info('开始合成HTML内容')
        html_template = '<table class="table"><thead><tr><th data-align="center">记录时间</th><th data-align="center">记录类型</th></tr></thead><tbody>{tbody}</tbody></table>'
        tbody = ['<tr><td>{0}</td><td>{1}</td></tr>'.format(record['created'], record['record_type']) for record in
                 records]
        html = html_template.format(tbody=''.join(tbody))
        result = {
            'error': False,
            'records': records,
            'person': personal_information,
            'html': html
        }
    except Person.DoesNotExist:
        result = {
            'error': True,
            'detail': '查询不到人员'
        }
        logger.exception('查询不到人员')
    except Exception as e:
        result = {
            'error': True,
            'detail': str(e)
        }
        logger.exception('获取数据失败')
    return JsonResponse(result)


def generate_excel_file(date):
    wb = Workbook()
    today = datetime.strptime(date, '%Y-%m-%d')
    attended_pks = Record.objects.filter(created__date=today).values_list('target', flat=True)

    attended_people = Person.objects.filter(Q(pk__in=attended_pks) & (Q(departure_date__gt=today) | Q(departure_date=None)) & Q(is_deleted=False))
    #attended_people = Person.objects.filter(Q(pk__in=attended_pks) & Q(is_active=True) & Q(is_deleted=False))
    attended_data = []
    for person in attended_people:
        starts = Record.objects.filter(Q(target=person) & Q(created__date=today) & Q(record_type='IN')).order_by(
            'created')
        start_time = starts.first().created.astimezone(tz) if starts.count() > 0 else '无记录'
        if start_time != '无记录':
            start = start_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            start = start_time
        ends = Record.objects.filter(Q(target=person) & Q(created__date=today) & Q(record_type='OUT')).order_by(
            '-created')
        end_time = ends.first().created.astimezone(tz) if ends.count() > 0 else '无记录'
        if end_time != '无记录':
            end = end_time.strftime("%Y-%m-%d %H:%M:%S")
        else:
            end = end_time
        
        attended_data.append({
            'pk': person.pk,
            'name': person.name,
            'position': person.position,
            'start': start,
            'end': end,
            'employee_number': person.employee_number,
            'department': person.department
        })

    ws = wb.create_sheet("出勤人员")
    ws.append(['姓名', '职位', '所属部门', '工号', '到达时间', '离开时间'])
    for person in attended_data:
        print('11111'+person['department'].name)
        ws.append([person['name'], person['position'], person['department'].name, person['employee_number'], person['start'],
                   person['end']])

    absent_people = Person.objects.filter(Q(is_deleted=False) & (Q(departure_date__gt=today) | Q(departure_date=None))).exclude(pk__in=attended_pks)
    ws = wb.create_sheet("缺勤人员")
    ws.append(['姓名', '所属部门', '职位', '工号'])
    for person in absent_people:
        ws.append([person.name, person.department.name, person.position, person.employee_number])

    records = Record.objects.filter(Q(created__date=today)).order_by('created')
    ws = wb.create_sheet("详细记录")
    ws.append(['姓名', '所属部门', '职位', '工号', '匹配时间', '记录类型'])
    for record in records:
        ws.append([record.target.name, record.target.department.name, record.target.position, record.target.employee_number,
                   record.created.astimezone(tz), record.record_type])

    wb.remove_sheet(wb.get_sheet_by_name('Sheet'))
    return wb


@require_http_methods(["GET"])
def export_test_records(request):
    date = request.GET.get('date')
    work_book = generate_excel_file(date)
    file_name = '{0}-{1}.xlsx'.format('出勤记录', date)
    response = HttpResponse(save_virtual_workbook(work_book), content_type='application/octet-stream')
    response['Content-Disposition'] = "attachment; filename*=utf-8''{}".format(escape_uri_path(file_name))
    return response
